import openpyxl


''' Lê a planilha sheet do arquivo file e escreve no script os trechos que criam
as tabelas. '''

def create_table (sheet, file):
    file.write ("CREATE TABLE %s(\n" % sheet.title.lower ())
    j = 1
    while (j < sheet.max_column):
        file.write ("    %s TEXT,\n" % (sheet.cell (1, j).value))
        j = j + 1
    file.write ("    %s TEXT);\n\n" % (sheet.cell (1, j).value))


''' Lê a planilha sheet do arquivo file e escreve no script os trechos que inserem
dados nas tabelas. '''

def write_insertions (sheet, file):
    i = 2
    while (i <= 20):
        file.write ("INSERT INTO %s (" % (sheet.title.lower ()))

        j = 1;
        while (j < sheet.max_column):
            file.write ("%s, " % (sheet.cell (1, j).value))
            j = j + 1
        file.write ("%s" % (sheet.cell (1, j).value))
        file.write (") ")

        file.write ("VALUES (")

        j = 1
        while (j < sheet.max_column):        
            file.write ("\"%s\", " % (sheet.cell (i,j).value))
            j = j + 1        
        file.write ("\"%s\"" % (sheet.cell (i,j).value))
        file.write (");\n")

        i = i + 1


def main ():
    file = open ("novo.sql", "x")
    file.write ("USE fakenews;\n\n")

    book = openpyxl.load_workbook ("[CEPI Fake News 2018] Base de Partes_v.2021.xlsx")
    sheet = book["Partes"]
    create_table (sheet, file)
    write_insertions (sheet, file)
    file.write ("\n")

    book = openpyxl.load_workbook ("[CEPI Fake News 2018] Base de Processos_v.2021.xlsx")
    sheet = book["Processos"]
    create_table (sheet, file)
    write_insertions (sheet, file)
    file.write ("\n")

    book = openpyxl.load_workbook ("[CEPI Fake News 2018] Base de Precedentes_v.2021.xlsx")
    sheet = book["precedentes_TSE"]
    create_table (sheet, file)
    write_insertions (sheet, file)
    file.write ("\n")

    sheet = book["precedentes_TREs"]
    create_table (sheet, file)
    write_insertions (sheet, file)
    file.write ("\n")

main ()
